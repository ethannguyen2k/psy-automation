import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import argparse
from pipeline import Pipeline

def create_test_excel():
    """Create a test Excel file with Wisemind Psychology example data."""
    wb = Workbook()
    ws = wb.active
    
    # Define headers
    headers = ["Practice", "Address", "Website", "Phone", "Name", "Email", "Doctors", "Type", 
              "Initial Consult", "Follow-up Consult", "Date", "Notes"]
    
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx).value = header
    
    # Real data for Wisemind Psychology
    test_data = [
        ["Wisemind Psychology", "Unit 2/40 Minchinton St, Caloundra QLD 4551", "http://www.wisemind.com.au", "0490193347", 
         "", "", "", "", "", "", "", ""]
    ]
    
    # Add data and highlight rows in green
    green_fill = PatternFill(start_color="FFA9D08E", end_color="FFA9D08E", fill_type="solid")
    
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            # Apply green fill to indicate the row should be processed
            if col_idx <= 4:  # Only highlight the first 4 columns
                cell.fill = green_fill
    
    # Save the test file
    test_file = "wisemind_test.xlsx"
    wb.save(test_file)
    print(f"Created test Excel file: {test_file}")
    return test_file

def create_sample_content_file():
    """Create a sample content file for Wisemind Psychology to simulate web scraping."""
    sample_dir = "wisemind_scraped_data"
    if not os.path.exists(sample_dir):
        os.makedirs(sample_dir)
        
    sample_content = """Practice: Wisemind Psychology
Website: http://www.wisemind.com.au
Emails: admin@wisemind.com.au
Doctor Pages: http://www.wisemind.com.au/our-team/

--- MAIN PAGE CONTENT ---

<h1>Welcome to Wisemind Psychology</h1>
<p>We are a team of dedicated psychologists providing mental health care and psychological services on the Sunshine Coast, Queensland.</p>
<p>Our clinic offers evidence-based psychological therapy for children, adolescents, and adults.</p>

<h2>Our Services</h2>
<ul>
<li>Individual therapy</li>
<li>Depression and anxiety treatment</li>
<li>Trauma therapy</li>
<li>Stress management</li>
<li>Relationship counseling</li>
</ul>

<h2>Contact Us</h2>
<p>Phone: 0490 193 347</p>
<p>Email: admin@wisemind.com.au</p>
<p>Address: Unit 2/40 Minchinton St, Caloundra QLD 4551</p>

--- DOCTOR PAGE: http://www.wisemind.com.au/our-team/ ---

<h1>Our Team</h1>
<p>Our clinic has a team of experienced and qualified psychologists.</p>

<h2>Melissa Madden</h2>
<p>Clinical Psychologist</p>
<p>Melissa is the principal psychologist at Wisemind Psychology with extensive experience in treating depression, anxiety, and trauma-related conditions.</p>

<h2>Danielle Comerford</h2>
<p>General Psychologist</p>
<p>Danielle specializes in working with adults experiencing anxiety, depression, and relationship difficulties.</p>

<h2>Rachel Hannam</h2>
<p>General Psychologist</p>
<p>Rachel works with children, adolescents, and adults, offering support for anxiety, depression, and stress management.</p>

--- OTHER PAGE: http://www.wisemind.com.au/fees/ ---

<h1>Fees and Rebates</h1>
<p>We aim to provide accessible psychological services to our community.</p>

<h2>Consultation Fees</h2>
<p>Initial Consultation (50 minutes): $220</p>
<p>Follow-up Sessions (50 minutes): $180</p>

<h2>Medicare Rebates</h2>
<p>With a Mental Health Treatment Plan from your GP, you may be eligible for Medicare rebates for psychology sessions.</p>

<h2>Private Health Insurance</h2>
<p>Clients with private health insurance may be eligible for rebates depending on their level of cover.</p>
"""
    
    sample_file = os.path.join(sample_dir, "Wisemind_Psychology_0.txt")
    with open(sample_file, 'w', encoding='utf-8') as f:
        f.write(sample_content)
        
    print(f"Created sample content file: {sample_file}")
    return sample_dir

def create_sample_extraction_result():
    """Create a sample extraction result file to simulate Gemini API output."""
    import json
    
    sample_dir = "wisemind_scraped_data"
    if not os.path.exists(sample_dir):
        os.makedirs(sample_dir)
        
    extraction_result = {
        "Wisemind_Psychology_0.txt": {
            "practice_name": "Wisemind Psychology",
            "email": "admin@wisemind.com.au",
            "doctor_page_url": "http://www.wisemind.com.au/our-team/",
            "psychologists": [
                {
                    "name": "Melissa Madden",
                    "type": "C"
                },
                {
                    "name": "Danielle Comerford",
                    "type": "G"
                },
                {
                    "name": "Rachel Hannam",
                    "type": "G"
                }
            ],
            "pricing_info": {
                "initial_consult": "$220",
                "followup_consult": "$180"
            }
        }
    }
    
    result_file = os.path.join(sample_dir, "extraction_results.json")
    with open(result_file, 'w', encoding='utf-8') as f:
        json.dump(extraction_result, f, indent=2)
        
    print(f"Created sample extraction result file: {result_file}")
    return result_file

def run_pipeline_test():
    """Run a test of the full pipeline with prepared data."""
    print("\n" + "="*50)
    print("TESTING FULL PIPELINE")
    print("="*50)
    
    # Check if API key is set
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("\nWarning: GEMINI_API_KEY not found in environment variables.")
        print("The API connection part of the pipeline may fail.")
        print("Set your API key with: export GEMINI_API_KEY='your-api-key'")
    
    # Create test directories
    output_dir = "pipeline_test_output"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        
    # Create test data
    test_file = create_test_excel()
    
    # Create sample content to simulate web scraping result
    scraped_data_dir = create_sample_content_file()
    
    # Create sample extraction result to simulate API output
    create_sample_extraction_result()
    
    # Configure and run the pipeline
    print("\nInitializing pipeline...")
    pipeline = Pipeline(
        input_file=test_file,
        output_dir=output_dir,
        scraped_data_dir=scraped_data_dir,
        batch_size=1,
        max_workers=1,
        api_key=api_key
    )
    
    # Test options:
    # 1. Run full pipeline - might fail if no API key or connection issues
    # 2. Run individual stages for testing
    
    # Option to select what to test
    test_mode = input("\nSelect test mode:\n1. Run full pipeline\n2. Run individual stages\nYour choice (1/2): ")
    
    if test_mode == "1":
        print("\nRunning full pipeline...")
        success = pipeline.run_pipeline()
        if success:
            print("\nFull pipeline test completed successfully!")
        else:
            print("\nFull pipeline test failed. Check logs for details.")
    else:
        print("\nRunning individual stages...")
        
        print("\nRunning Stage 1: Excel parsing and initial validation")
        success = pipeline.run_stage1()
        print(f"Stage 1 {'completed successfully' if success else 'failed'}")
        
        print("\nRunning Stage 4: Validation and structural formatting")
        # We need to set up the extracted data since we're skipping Stage 2-3
        import json
        with open(os.path.join(scraped_data_dir, "extraction_results.json"), 'r') as f:
            pipeline.extracted_data = json.load(f)
        success = pipeline.run_stage4()
        print(f"Stage 4 {'completed successfully' if success else 'failed'}")
        
        print("\nRunning Stage 5: Excel output generation")
        success = pipeline.run_stage5()
        print(f"Stage 5 {'completed successfully' if success else 'failed'}")
        
        print("\nIndividual stages test completed!")
    
    print(f"\nCheck {output_dir} directory for output files")

if __name__ == "__main__":
    run_pipeline_test()