import pandas as pd
import os
import logging
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("log/excel_output.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class ExcelOutputGenerator:
    def __init__(self, output_file=None):
        """
        Initialize the Excel output generator.
        
        Args:
            output_file (str): Path to save the output Excel file
        """
        self.output_file = output_file
        
        # Define fill styles
        self.green_fill = PatternFill(start_color="FFA9D08E", end_color="FFA9D08E", fill_type="solid")
        self.yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        self.light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    
    def generate_excel(self, df, output_file=None, green_rows=None):
        """
        Generate Excel file from DataFrame.
        
        Args:
            df (pandas.DataFrame): DataFrame to save
            output_file (str): Path to save the output Excel file
            green_rows (list): List of row indices that were originally green
            
        Returns:
            str: Path to saved Excel file
        """
        if output_file:
            self.output_file = output_file
            
        if not self.output_file:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            self.output_file = f"psychology_clinics_processed_{timestamp}.xlsx"
        
        # If green_rows not provided, assume first rows up to any additional rows
        if green_rows is None:
            # Estimate original green rows (this is a rough guess)
            green_rows = list(range(len(df)))
            
        # Format phone numbers as text to preserve formatting
        df_formatted = df.copy()
        if 'Phone' in df_formatted.columns:
            # Convert phone numbers to strings with formatting preserved
            df_formatted['Phone'] = df_formatted['Phone'].astype(str)
        
        # Save DataFrame to Excel with phone numbers as text
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            df_formatted.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            
            # Format phone column as text
            if 'Phone' in df_formatted.columns:
                phone_col_idx = list(df_formatted.columns).index('Phone') + 1  # +1 because openpyxl is 1-indexed
                for row_idx in range(2, len(df_formatted) + 2):  # +2 to account for header and 1-indexing
                    cell = worksheet.cell(row=row_idx, column=phone_col_idx)
                    cell.number_format = '@'  # Format as text
                    
        logger.info(f"Saved data to {self.output_file}")
        
        # Apply formatting
        self._apply_formatting(green_rows)
        
        return self.output_file
    
    def _apply_formatting(self, green_rows):
        """Apply formatting to the Excel file."""
        try:
            wb = load_workbook(self.output_file)
            ws = wb.active
            
            # Format header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            header_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
                
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
            
            # Format data rows (auto-width + borders)
            data_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = data_border
                    
                    # Apply alignment based on cell content
                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Apply colored fills for specific content
            for row_idx in range(2, ws.max_row + 1):
                # Apply green fill for rows that were initially green
                for col_idx in range(1, 5):  # First 4 columns
                    ws.cell(row=row_idx, column=col_idx).fill = self.green_fill
                
                # Apply yellow fill for cells with potential issues
                notes_cell = ws.cell(row=row_idx, column=ws.max_column)
                if notes_cell.value and "discrepancy" in str(notes_cell.value).lower():
                    notes_cell.fill = self.yellow_fill
                    
                # Apply light blue fill for new psychologist rows
                if row_idx > len(green_rows) + 1:  # +1 for header row
                    ws.cell(row=row_idx, column=5).fill = self.light_blue_fill  # Name column
            
            # Save the workbook
            wb.save(self.output_file)
            
        except Exception as e:
            logger.error(f"Error applying formatting: {str(e)}")