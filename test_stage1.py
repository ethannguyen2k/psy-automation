import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from stage1_excel_parsing import ExcelProcessor

def create_test_excel():
    """Create a test Excel file with Wisemind Psychology example data."""
    wb = Workbook()
    ws = wb.active
    
    # Define headers
    headers = ["Practice", "Address", "Website", "Phone", "Name", "Email", "Doctors", "Type", 
              "Initial Consult", "Follow-up Consult", "Date", "Notes"]
    
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx).value = header
    
    # Real data for Wisemind Psychology
    test_data = [
        ["Wisemind Psychology", "Unit 2/40 Minchinton St, Caloundra QLD 4551", "http://www.wisemind.com.au", "0490193347", 
         "", "", "", "", "", "", "", ""]
    ]
    
    # Add data and highlight rows in green
    green_fill = PatternFill(start_color="FFA9D08E", end_color="FFA9D08E", fill_type="solid")
    
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            # Apply green fill to indicate the row should be processed
            if col_idx <= 4:  # Only highlight the first 4 columns
                cell.fill = green_fill
    
    # Save the test file
    test_file = "test_clinics.xlsx"
    wb.save(test_file)
    print(f"Created test Excel file: {test_file}")
    return test_file

def run_test():
    """Run the test for the ExcelProcessor class using Wisemind Psychology example."""
    # Create a test Excel file
    test_file = create_test_excel()
    
    print("\n" + "="*50)
    print("TESTING EXCEL PROCESSING WITH WISEMIND PSYCHOLOGY")
    print("="*50 + "\n")
    
    # Initialize and run the processor
    processor = ExcelProcessor(test_file)
    validation_results = processor.run_initial_validation()
    
    # Print validation results
    print("\nValidation Results Summary:")
    print(f"Address Format Issues: {len(validation_results['address_format_issues']) if isinstance(validation_results['address_format_issues'], list) else validation_results['address_format_issues']}")
    print(f"Duplicate Phone Numbers: {len(validation_results['duplicate_phones'])}")
    print(f"Rows with Missing Data: {len(validation_results['missing_data'])}")
    print(f"Invalid Website URLs: {len(validation_results['invalid_websites'])}")
    
    # Show detailed results for Wisemind Psychology
    print("\nDetailed Results for Wisemind Psychology:")
    
    # Check website format (should identify as missing http:// prefix)
    website = "http://www.wisemind.com.au"
    print(f"Website: {website}")
    print(f"Website format valid: {processor.is_valid_url(website) if hasattr(processor, 'is_valid_url') else 'N/A'}")
    
    # Check address format
    address = "Unit 2/40 Minchinton St, Caloundra QLD 4551"
    print(f"Address: {address}")
    address_pattern1 = r"^\d+\s+[\w\s]+,\s+[\w\s]+\s+[A-Z]{2,3}\s+\d{4,5}$"
    address_pattern2 = r"^Cnr\s+[\w\s]+\s+&\s+[\w\s]+,\s+[\w\s]+\s+[A-Z]{2,3}\s+\d{4,5}$"
    import re
    print(f"Address matches standard pattern: {bool(re.match(address_pattern1, address))}")
    print(f"Address matches corner pattern: {bool(re.match(address_pattern2, address))}")
    
    # Save the processed file
    processor.save_results()
    
    print("\nTest completed. Check the generated files for results.")

if __name__ == "__main__":
    run_test()