import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class ExcelProcessor:
    def __init__(self, file_path):
        """
        Initialize the Excel processor with the path to the excel file.
        
        Args:
            file_path (str): Path to the excel file
        """
        self.file_path = file_path
        self.df = None
        self.green_rows = None
        self.workbook = None
        self.output_path = self._generate_output_path(file_path)
        
    def _generate_output_path(self, input_path):
        """Generate output path with '_processed' suffix."""
        base, ext = os.path.splitext(input_path)
        return f"{base}_processed{ext}"
    
    def load_excel(self):
        """Load the excel file and extract the data."""
        try:
            # Load the excel file with pandas to work with the data
            self.df = pd.read_excel(self.file_path)
            
            # Also load with openpyxl to preserve formatting
            self.workbook = load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            
            print(f"Successfully loaded {self.file_path}")
            print(f"Columns found: {list(self.df.columns)}")
            return True
        except Exception as e:
            print(f"Error loading excel file: {e}")
            return False
    
    def identify_green_rows(self):
        """
        Identify rows that are highlighted in green.
        These are the rows that need to be processed.
        """
        green_rows = []
        
        # Iterate through the rows to find green cells
        for row_idx in range(2, self.worksheet.max_row + 1):  # Skip header row
            if self._is_row_green(row_idx):
                green_rows.append(row_idx - 2)  # Adjust for pandas 0-indexing
        
        self.green_rows = green_rows
        print(f"Found {len(green_rows)} green rows to process")
        return green_rows
    
    def _is_row_green(self, row_idx):
        """Check if a row is highlighted in green."""
        # Check any cell in the row to see if it's green
        # Green fill typically has an rgb value close to (144, 238, 144)
        # We'll check the first cell in each row
        cell = self.worksheet.cell(row=row_idx, column=1)
        if cell.fill.start_color.rgb in ['FFA9D08E','FFA8D08D']:
            return True
        
        # Check if any cell in the row has a green fill
        for col_idx in range(1, 5):  # Check first 4 columns
            cell = self.worksheet.cell(row=row_idx, column=col_idx)
            if cell.fill.start_color.rgb in ['FFA9D08E','FFA8D08D']:
                return True
        
        return False
    
    def validate_address_format(self):
        """Validate and correct the address format according to requirements."""
        if self.df is None or self.green_rows is None:
            print("Excel file not loaded or green rows not identified")
            return
        
        # Address format patterns
        address_pattern1 = r"^\d+\s+[\w\s]+,\s+[\w\s]+\s+[A-Z]{2,3}\s+\d{4,5}$"  # Pattern for standard addresses
        address_pattern2 = r"^Cnr\s+[\w\s]+\s+&\s+[\w\s]+,\s+[\w\s]+\s+[A-Z]{2,3}\s+\d{4,5}$"  # Pattern for corner addresses
        
        format_issues = 0
        corrected_addresses = 0
        
        # Process only green rows
        for idx in self.green_rows:
            if idx >= len(self.df):
                continue
                
            row = self.df.iloc[idx]
            
            # Check address format
            if 'Address' in row and pd.notna(row['Address']):
                address = str(row['Address']).strip()
                original_address = address
                
                # Improved unit/suite pattern matching
                # Look for patterns like "Unit 2/40", "2/7", "35B/12", etc. at the beginning of addresses
                unit_patterns = [
                    r"^(Unit\s+)?(\d+[A-Za-z]?)[/\\](\d+)\s+([\w\s]+)",  # Unit X/Y Street or X/Y Street
                    r"^Suite\s+(\d+)[/\\](\d+)\s+([\w\s]+)",  # Suite X/Y Street
                    r"^Shop\s+(\d+)[/\\](\d+)\s+([\w\s]+)"    # Shop X/Y Street
                ]
                
                for pattern in unit_patterns:
                    match = re.search(pattern, address)
                    if match:
                        # Extract building number and street
                        if pattern.startswith("^(Unit"):
                            # For pattern matching "Unit X/Y" or "X/Y"
                            building_number = match.group(3)  # The Y in X/Y
                            street = match.group(4)
                        else:
                            # For other patterns
                            building_number = match.group(2)  # The Y in X/Y
                            street = match.group(3)
                        
                        # Extract everything after the street name portion
                        rest_parts = address.split(street, 1)
                        if len(rest_parts) > 1:
                            rest = rest_parts[1]
                        else:
                            rest = ""
                        
                        # Reconstruct address with building number only (no unit)
                        address = f"{building_number} {street}{rest}"
                        break
                
                # Check if address matches either pattern after cleaning
                if not (re.match(address_pattern1, address) or re.match(address_pattern2, address)):
                    print(f"Row {idx+2}: Address format issue - {address}")
                    format_issues += 1
                elif address != original_address:
                    # Update the address in the DataFrame
                    self.df.at[idx, 'Address'] = address
                    print(f"Row {idx+2}: Address corrected from '{original_address}' to '{address}'")
                    corrected_addresses += 1
        
        print(f"Found {format_issues} address format issues")
        print(f"Corrected {corrected_addresses} addresses")
        return format_issues
    
    def check_phone_duplicates(self):
        """Check for duplicate phone numbers and flag them."""
        if self.df is None or self.green_rows is None:
            print("Excel file not loaded or green rows not identified")
            return
        
        phone_counts = {}
        duplicate_phones = []
        
        # Count occurrences of each phone number
        for idx in self.green_rows:
            if idx >= len(self.df):
                continue
                
            row = self.df.iloc[idx]
            
            if 'Phone' in row and pd.notna(row['Phone']):
                phone = str(row['Phone']).strip()
                if phone in phone_counts:
                    phone_counts[phone].append(idx)
                else:
                    phone_counts[phone] = [idx]
        
        # Identify duplicates
        for phone, indices in phone_counts.items():
            if len(indices) > 1:
                duplicate_phones.append((phone, indices))
        
        print(f"Found {len(duplicate_phones)} duplicate phone numbers")
        return duplicate_phones
    
    def check_missing_data(self):
        """Check for missing addresses or phone numbers."""
        if self.df is None or self.green_rows is None:
            print("Excel file not loaded or green rows not identified")
            return
        
        missing_data = []
        
        for idx in self.green_rows:
            if idx >= len(self.df):
                continue
                
            row = self.df.iloc[idx]
            
            missing_fields = []
            if 'Address' in row and (pd.isna(row['Address']) or str(row['Address']).strip() == ""):
                missing_fields.append('Address')
            
            if 'Phone' in row and (pd.isna(row['Phone']) or str(row['Phone']).strip() == ""):
                missing_fields.append('Phone')
            
            if missing_fields:
                missing_data.append((idx, missing_fields))
        
        print(f"Found {len(missing_data)} rows with missing critical data")
        return missing_data
    
    def validate_websites(self):
        """Validate website URLs."""
        if self.df is None or self.green_rows is None:
            print("Excel file not loaded or green rows not identified")
            return
        
        invalid_websites = []
        
        # Basic URL pattern for validation
        url_pattern = r'^(http|https)://[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}(/.*)?$'
        
        for idx in self.green_rows:
            if idx >= len(self.df):
                continue
                
            row = self.df.iloc[idx]
            
            if 'Website' in row and pd.notna(row['Website']):
                website = str(row['Website']).strip()
                if not re.match(url_pattern, website):
                    invalid_websites.append((idx, website))
        
        print(f"Found {len(invalid_websites)} invalid website URLs")
        return invalid_websites
    
    def run_initial_validation(self):
        """Run all validation checks and create a summary report."""
        if not self.load_excel():
            return "Failed to load Excel file"
        
        self.identify_green_rows()
        
        validation_report = {
            "address_format_issues": self.validate_address_format(),
            "duplicate_phones": self.check_phone_duplicates(),
            "missing_data": self.check_missing_data(),
            "invalid_websites": self.validate_websites()
        }
        
        return validation_report
    
    def save_results(self):
        """Save the processed DataFrame back to an Excel file."""
        try:
            self.df.to_excel(self.output_path, index=False)
            print(f"Saved processed data to {self.output_path}")
            return True
        except Exception as e:
            print(f"Error saving processed data: {e}")
            return False

# Example usage
if __name__ == "__main__":
    processor = ExcelProcessor("3 NT Research Outgoing 1.xlsx")
    validation_results = processor.run_initial_validation()
    print("\nValidation Results Summary:")
    print(f"Address Format Issues: {len(validation_results['address_format_issues']) if isinstance(validation_results['address_format_issues'], list) else validation_results['address_format_issues']}")
    print(f"Duplicate Phone Numbers: {len(validation_results['duplicate_phones'])}")
    print(f"Rows with Missing Data: {len(validation_results['missing_data'])}")
    print(f"Invalid Website URLs: {len(validation_results['invalid_websites'])}")
    
    # Save the processed file
    processor.save_results()